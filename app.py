"""
Backend API para Sistema de Gestión de Tareas
FastAPI + SQLite
Versión 2.0 - Con comentarios en registros de tiempo
"""

from fastapi import FastAPI, HTTPException, Query
from fastapi.responses import FileResponse
from fastapi.middleware.cors import CORSMiddleware
from pydantic import BaseModel
from typing import Optional
import sqlite3
from datetime import datetime
import os

app = FastAPI(
    title="Inmotica TaskFlow API",
    description="Sistema de Gestión y Control de Tareas",
    version="2.0.0"
)

# Configurar CORS
app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

DATABASE = 'Inmotica-tasks.db'

# ==================== MODELOS PYDANTIC ====================

class UserCreate(BaseModel):
    name: str
    email: Optional[str] = None

class TaskCreate(BaseModel):
    name: str
    description: Optional[str] = None
    user_id: int
    max_time_minutes: Optional[int] = 0
    max_date: Optional[str] = None
    status: Optional[str] = "Pendiente"

class TaskUpdate(BaseModel):
    name: str
    description: Optional[str] = None
    user_id: int
    max_time_minutes: Optional[int] = 0
    max_date: Optional[str] = None
    status: Optional[str] = "Pendiente"

class AnnotationCreate(BaseModel):
    text: str

class AnnotationUpdate(BaseModel):
    text: str

class TimeEntryCreate(BaseModel):
    start_time: str
    end_time: Optional[str] = None
    comment: Optional[str] = None

class TimeEntryUpdate(BaseModel):
    start_time: str
    end_time: Optional[str] = None
    comment: Optional[str] = None

# ==================== DATABASE ====================

def get_db():
    """Obtener conexión a la base de datos"""
    conn = sqlite3.connect(DATABASE)
    conn.row_factory = sqlite3.Row
    return conn

def init_db():
    """Inicializar la base de datos con las tablas necesarias"""
    conn = get_db()
    cursor = conn.cursor()
    
    # Tabla de usuarios
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS users (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            name TEXT NOT NULL,
            email TEXT,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        )
    ''')
    
    # Tabla de tareas
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS tasks (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            task_number INTEGER UNIQUE NOT NULL,
            name TEXT NOT NULL,
            description TEXT,
            user_id INTEGER NOT NULL,
            max_time_minutes INTEGER DEFAULT 0,
            max_date DATE,
            status TEXT DEFAULT 'Pendiente',
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            FOREIGN KEY (user_id) REFERENCES users(id)
        )
    ''')
    
    # Migración: Añadir campo status si no existe
    try:
        cursor.execute("SELECT status FROM tasks LIMIT 1")
    except sqlite3.OperationalError:
        print("🔄 Migrando: añadiendo campo 'status' a tabla tasks...")
        cursor.execute("ALTER TABLE tasks ADD COLUMN status TEXT DEFAULT 'Pendiente'")
        cursor.execute("UPDATE tasks SET status = 'Pendiente' WHERE status IS NULL")
        conn.commit()
        print("✅ Campo 'status' añadido")
    
    # Tabla de anotaciones
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS annotations (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            task_id INTEGER NOT NULL,
            text TEXT NOT NULL,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            FOREIGN KEY (task_id) REFERENCES tasks(id) ON DELETE CASCADE
        )
    ''')
    
    # Tabla de registros de tiempo
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS time_entries (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            task_id INTEGER NOT NULL,
            start_time TIMESTAMP NOT NULL,
            end_time TIMESTAMP,
            duration_minutes INTEGER,
            comment TEXT,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            FOREIGN KEY (task_id) REFERENCES tasks(id) ON DELETE CASCADE
        )
    ''')
    
    # Migración: Añadir campo comment si no existe
    try:
        cursor.execute("SELECT comment FROM time_entries LIMIT 1")
    except sqlite3.OperationalError:
        print("🔄 Migrando: añadiendo campo 'comment' a time_entries...")
        cursor.execute("ALTER TABLE time_entries ADD COLUMN comment TEXT")
        conn.commit()
        print("✅ Campo 'comment' añadido")
    
    conn.commit()
    conn.close()

def calculate_duration(start_time, end_time):
    """Calcular duración en minutos entre dos fechas"""
    if not end_time:
        return None
    
    start = datetime.fromisoformat(start_time.replace('Z', '+00:00'))
    end = datetime.fromisoformat(end_time.replace('Z', '+00:00'))
    duration = (end - start).total_seconds() / 60
    return int(duration)

# ==================== EVENTOS ====================

@app.on_event("startup")
async def startup_event():
    """Inicializar base de datos al arrancar"""
    init_db()
    print("✅ FastAPI: Base de datos inicializada")
    print("📝 Documentación: http://localhost:5000/docs")

@app.get("/")
async def root():
    """Endpoint raíz"""
    return {
        "message": "Inmotica TaskFlow API v2.0",
        "framework": "FastAPI",
        "docs": "/docs",
        "redoc": "/redoc"
    }

# ==================== USUARIOS ====================

@app.get('/api/users')
async def get_users():
    """Obtener todos los usuarios"""
    conn = get_db()
    users = conn.execute('SELECT * FROM users ORDER BY name').fetchall()
    conn.close()
    return [dict(user) for user in users]

@app.post('/api/users', status_code=201)
async def create_user(user: UserCreate):
    """Crear nuevo usuario"""
    conn = get_db()
    cursor = conn.cursor()
    
    cursor.execute(
        'INSERT INTO users (name, email) VALUES (?, ?)',
        (user.name, user.email)
    )
    
    conn.commit()
    user_id = cursor.lastrowid
    conn.close()
    
    return {'id': user_id, 'message': 'Usuario creado'}

@app.delete('/api/users/{user_id}')
async def delete_user(user_id: int):
    """Eliminar usuario"""
    conn = get_db()
    conn.execute('DELETE FROM users WHERE id = ?', (user_id,))
    conn.commit()
    conn.close()
    return {'message': 'Usuario eliminado'}

# ==================== TAREAS ====================

@app.get('/api/tasks')
async def get_tasks():
    """Obtener todas las tareas"""
    conn = get_db()
    tasks = conn.execute('''
        SELECT t.*, u.name as user_name
        FROM tasks t
        LEFT JOIN users u ON t.user_id = u.id
        ORDER BY t.task_number
    ''').fetchall()
    conn.close()
    return [dict(task) for task in tasks]

@app.post('/api/tasks', status_code=201)
async def create_task(task: TaskCreate):
    """Crear nueva tarea"""
    conn = get_db()
    cursor = conn.cursor()
    
    # Obtener el siguiente número de tarea
    last_task = cursor.execute('SELECT MAX(task_number) as max_num FROM tasks').fetchone()
    next_number = (last_task['max_num'] or 0) + 1
    
    cursor.execute('''
        INSERT INTO tasks (task_number, name, description, user_id, max_time_minutes, max_date, status)
        VALUES (?, ?, ?, ?, ?, ?, ?)
    ''', (
        next_number,
        task.name,
        task.description,
        task.user_id,
        task.max_time_minutes,
        task.max_date,
        task.status
    ))
    
    conn.commit()
    task_id = cursor.lastrowid
    conn.close()
    
    return {'id': task_id, 'task_number': next_number, 'message': 'Tarea creada'}

@app.put('/api/tasks/{task_id}')
async def update_task(task_id: int, task: TaskUpdate):
    """Actualizar tarea"""
    conn = get_db()
    cursor = conn.cursor()
    
    cursor.execute('''
        UPDATE tasks 
        SET name = ?, description = ?, user_id = ?, max_time_minutes = ?, max_date = ?, status = ?
        WHERE id = ?
    ''', (
        task.name,
        task.description,
        task.user_id,
        task.max_time_minutes,
        task.max_date,
        task.status,
        task_id
    ))
    
    conn.commit()
    conn.close()
    
    return {'id': task_id, 'message': 'Tarea actualizada'}

@app.delete('/api/tasks/{task_id}')
async def delete_task(task_id: int):
    """Eliminar tarea"""
    conn = get_db()
    conn.execute('DELETE FROM tasks WHERE id = ?', (task_id,))
    conn.commit()
    conn.close()
    return {'message': 'Tarea eliminada'}

# ==================== ANOTACIONES ====================

@app.get('/api/tasks/{task_id}/annotations')
async def get_annotations(task_id: int):
    """Obtener anotaciones de una tarea"""
    conn = get_db()
    annotations = conn.execute('''
        SELECT * FROM annotations 
        WHERE task_id = ? 
        ORDER BY created_at DESC
    ''', (task_id,)).fetchall()
    conn.close()
    return [dict(annotation) for annotation in annotations]

@app.post('/api/tasks/{task_id}/annotations', status_code=201)
async def create_annotation(task_id: int, annotation: AnnotationCreate):
    """Crear nueva anotación"""
    conn = get_db()
    cursor = conn.cursor()
    
    cursor.execute(
        'INSERT INTO annotations (task_id, text) VALUES (?, ?)',
        (task_id, annotation.text)
    )
    
    conn.commit()
    annotation_id = cursor.lastrowid
    conn.close()
    
    return {'id': annotation_id, 'message': 'Anotación creada'}

@app.put('/api/annotations/{annotation_id}')
async def update_annotation(annotation_id: int, annotation: AnnotationUpdate):
    """Actualizar anotación"""
    conn = get_db()
    cursor = conn.cursor()
    
    cursor.execute(
        'UPDATE annotations SET text = ? WHERE id = ?',
        (annotation.text, annotation_id)
    )
    
    conn.commit()
    conn.close()
    
    return {'id': annotation_id, 'message': 'Anotación actualizada'}

@app.delete('/api/annotations/{annotation_id}')
async def delete_annotation(annotation_id: int):
    """Eliminar anotación"""
    conn = get_db()
    conn.execute('DELETE FROM annotations WHERE id = ?', (annotation_id,))
    conn.commit()
    conn.close()
    return {'message': 'Anotación eliminada'}

# ==================== REGISTROS DE TIEMPO ====================

@app.get('/api/tasks/{task_id}/times')
async def get_time_entries(task_id: int):
    """Obtener registros de tiempo de una tarea"""
    conn = get_db()
    times = conn.execute('''
        SELECT * FROM time_entries 
        WHERE task_id = ? 
        ORDER BY start_time DESC
    ''', (task_id,)).fetchall()
    conn.close()
    return [dict(time) for time in times]

@app.post('/api/tasks/{task_id}/times', status_code=201)
async def create_time_entry(task_id: int, time_entry: TimeEntryCreate):
    """Crear nuevo registro de tiempo con comentario"""
    conn = get_db()
    cursor = conn.cursor()
    
    duration = None
    if time_entry.end_time:
        duration = calculate_duration(time_entry.start_time, time_entry.end_time)
    
    cursor.execute(
        'INSERT INTO time_entries (task_id, start_time, end_time, duration_minutes, comment) VALUES (?, ?, ?, ?, ?)',
        (task_id, time_entry.start_time, time_entry.end_time, duration, time_entry.comment)
    )
    
    conn.commit()
    entry_id = cursor.lastrowid
    conn.close()
    
    return {'id': entry_id, 'message': 'Registro creado', 'duration_minutes': duration}

@app.put('/api/times/{time_id}')
async def update_time_entry(time_id: int, time_entry: TimeEntryUpdate):
    """Actualizar registro de tiempo"""
    conn = get_db()
    cursor = conn.cursor()
    
    duration = None
    if time_entry.end_time:
        duration = calculate_duration(time_entry.start_time, time_entry.end_time)
        if duration < 0:
            conn.close()
            raise HTTPException(status_code=400, detail='La fecha de fin debe ser posterior a la de inicio')
    
    cursor.execute(
        'UPDATE time_entries SET start_time = ?, end_time = ?, duration_minutes = ?, comment = ? WHERE id = ?',
        (time_entry.start_time, time_entry.end_time, duration, time_entry.comment, time_id)
    )
    
    conn.commit()
    conn.close()
    
    return {'id': time_id, 'message': 'Registro actualizado', 'duration_minutes': duration}

@app.delete('/api/times/{time_id}')
async def delete_time_entry(time_id: int):
    """Eliminar registro de tiempo"""
    conn = get_db()
    conn.execute('DELETE FROM time_entries WHERE id = ?', (time_id,))
    conn.commit()
    conn.close()
    return {'message': 'Registro eliminado'}

# ==================== MAIN ====================

if __name__ == "__main__":
    import uvicorn
    uvicorn.run(app, host="0.0.0.0", port=5000)

# ==================== INFORMES ====================

@app.get('/api/reports/excel')
async def generate_excel_report(
    from_task: int = Query(..., alias='from'),
    to_task: int = Query(..., alias='to'),
    user_id: Optional[int] = None,
    status: Optional[str] = None
):
    """Generar informe en Excel por rango de tareas"""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
        
        conn = get_db()
        
        # Query con filtro opcional de usuario y estado
        query = '''
            SELECT t.*, u.name as user_name
            FROM tasks t
            LEFT JOIN users u ON t.user_id = u.id
            WHERE t.task_number BETWEEN ? AND ?
        '''
        params = [from_task, to_task]
        
        if user_id:
            query += ' AND t.user_id = ?'
            params.append(int(user_id))
        
        if status:
            query += ' AND t.status = ?'
            params.append(status)
        
        query += ' ORDER BY t.task_number'
        
        tasks = conn.execute(query, params).fetchall()
        
        if not tasks:
            conn.close()
            raise HTTPException(status_code=404, detail='No se encontraron tareas en ese rango')
        
        wb = Workbook()
        wb.remove(wb.active)
        
        # Crear una hoja por cada tarea
        for task in tasks:
            task_dict = dict(task)
            sheet_name = f"{task_dict['task_number']} - {task_dict['name'][:25]}"
            ws = wb.create_sheet(sheet_name)
            
            # Título
            ws['A1'] = f"Tarea #{task_dict['task_number']}: {task_dict['name']}"
            ws['A1'].font = Font(size=14, bold=True, color='EF8354')
            ws.merge_cells('A1:E1')
            
            # Información de la tarea
            ws['A2'] = f"Asignado a: {task_dict['user_name']}"
            ws['A3'] = f"Estado: {task_dict['status']}"
            ws['A4'] = f"Tiempo máximo: {task_dict['max_time_minutes']} minutos"
            if task_dict['max_date']:
                ws['A5'] = f"Fecha límite: {task_dict['max_date']}"
            
            if task_dict['description']:
                ws['A6'] = f"Descripción: {task_dict['description']}"
            
            # Encabezados de registros de tiempo
            row = 8
            ws[f'A{row}'] = 'Fecha/Hora Inicio'
            ws[f'B{row}'] = 'Fecha/Hora Fin'
            ws[f'C{row}'] = 'Duración (minutos)'
            ws[f'D{row}'] = 'Comentario'
            
            for col in ['A', 'B', 'C', 'D']:
                cell = ws[f'{col}{row}']
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color='4F5D75', end_color='4F5D75', fill_type='solid')
                cell.alignment = Alignment(horizontal='center')
            
            # Obtener registros de tiempo
            times = conn.execute('''
                SELECT * FROM time_entries
                WHERE task_id = ?
                ORDER BY start_time
            ''', (task_dict['id'],)).fetchall()
            
            row += 1
            total_minutes = 0
            
            for time_entry in times:
                time_dict = dict(time_entry)
                ws[f'A{row}'] = time_dict['start_time']
                ws[f'B{row}'] = time_dict['end_time'] if time_dict['end_time'] else 'En progreso'
                
                if time_dict['duration_minutes']:
                    ws[f'C{row}'] = time_dict['duration_minutes']
                    total_minutes += time_dict['duration_minutes']
                else:
                    ws[f'C{row}'] = '-'
                
                ws[f'D{row}'] = time_dict['comment'] if time_dict['comment'] else '-'
                
                row += 1
            
            # Total
            row += 1
            ws[f'A{row}'] = 'TOTAL'
            ws[f'A{row}'].font = Font(bold=True)
            ws[f'C{row}'] = total_minutes
            ws[f'C{row}'].font = Font(bold=True)
            
            # Ajustar anchos
            ws.column_dimensions['A'].width = 20
            ws.column_dimensions['B'].width = 20
            ws.column_dimensions['C'].width = 20
            ws.column_dimensions['D'].width = 40
        
        conn.close()
        
        # Guardar archivo
        user_suffix = f'_usuario{user_id}' if user_id else ''
        status_suffix = f'_estado{status.replace(" ", "")}' if status else ''
        filename = f'informe_tareas_{from_task}-{to_task}{user_suffix}{status_suffix}.xlsx'
        wb.save(filename)
        
        return FileResponse(
            filename, 
            filename=filename,
            media_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        
    except Exception as e:
        print(f"Error generando Excel: {e}")
        raise HTTPException(status_code=500, detail=str(e))

@app.get('/api/reports/pdf')
async def generate_pdf_report(
    from_task: int = Query(..., alias='from'),
    to_task: int = Query(..., alias='to'),
    user_id: Optional[int] = None,
    status: Optional[str] = None
):
    """Generar informe en PDF por rango de tareas"""
    try:
        from reportlab.lib.pagesizes import letter
        from reportlab.lib import colors
        from reportlab.lib.units import inch
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.lib.enums import TA_CENTER
        
        conn = get_db()
        
        query = '''
            SELECT t.*, u.name as user_name
            FROM tasks t
            LEFT JOIN users u ON t.user_id = u.id
            WHERE t.task_number BETWEEN ? AND ?
        '''
        params = [from_task, to_task]
        
        if user_id:
            query += ' AND t.user_id = ?'
            params.append(int(user_id))
        
        if status:
            query += ' AND t.status = ?'
            params.append(status)
        
        query += ' ORDER BY t.task_number'
        
        tasks = conn.execute(query, params).fetchall()
        
        if not tasks:
            conn.close()
            raise HTTPException(status_code=404, detail='No se encontraron tareas en ese rango')
        
        user_suffix = f'_usuario{user_id}' if user_id else ''
        status_suffix = f'_estado{status.replace(" ", "")}' if status else ''
        filename = f'informe_tareas_{from_task}-{to_task}{user_suffix}{status_suffix}.pdf'
        doc = SimpleDocTemplate(filename, pagesize=letter)
        story = []
        styles = getSampleStyleSheet()
        
        # Estilos personalizados
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=16,
            textColor=colors.HexColor('#EF8354'),
            spaceAfter=12,
            alignment=TA_CENTER
        )
        
        heading_style = ParagraphStyle(
            'CustomHeading',
            parent=styles['Heading2'],
            fontSize=14,
            textColor=colors.HexColor('#4F5D75'),
            spaceAfter=10
        )
        
        # Título principal
        story.append(Paragraph('INFORME DE TAREAS', title_style))
        story.append(Paragraph(f'Tareas #{from_task} a #{to_task}', styles['Normal']))
        story.append(Spacer(1, 0.3*inch))
        
        # Procesar cada tarea
        for task in tasks:
            task_dict = dict(task)
            
            # Título de tarea
            story.append(Paragraph(f"Tarea #{task_dict['task_number']}: {task_dict['name']}", heading_style))
            
            # Información de la tarea
            info_text = f"<b>Asignado a:</b> {task_dict['user_name']}<br/>"
            info_text += f"<b>Estado:</b> {task_dict['status']}<br/>"
            info_text += f"<b>Tiempo máximo:</b> {task_dict['max_time_minutes']} minutos<br/>"
            if task_dict['max_date']:
                info_text += f"<b>Fecha límite:</b> {task_dict['max_date']}<br/>"
            if task_dict['description']:
                info_text += f"<b>Descripción:</b> {task_dict['description']}<br/>"
            
            story.append(Paragraph(info_text, styles['Normal']))
            story.append(Spacer(1, 0.2*inch))
            
            # Obtener registros de tiempo
            times = conn.execute('''
                SELECT * FROM time_entries
                WHERE task_id = ?
                ORDER BY start_time
            ''', (task_dict['id'],)).fetchall()
            
            if times:
                # Tabla de tiempos con comentarios
                data = [['Inicio', 'Fin', 'Duración', 'Comentario']]
                total_minutes = 0
                
                for time_entry in times:
                    time_dict = dict(time_entry)
                    data.append([
                        time_dict['start_time'],
                        time_dict['end_time'] if time_dict['end_time'] else 'En progreso',
                        str(time_dict['duration_minutes']) if time_dict['duration_minutes'] else '-',
                        time_dict['comment'] if time_dict['comment'] else '-'
                    ])
                    
                    if time_dict['duration_minutes']:
                        total_minutes += time_dict['duration_minutes']
                
                # Fila de total
                if total_minutes > 0:
                    data.append(['', '', str(total_minutes), 'TOTAL'])
                
                table = Table(data, colWidths=[1.5*inch, 1.5*inch, 1*inch, 3*inch])
                table.setStyle(TableStyle([
                    ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#4F5D75')),
                    ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                    ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                    ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                    ('FONTSIZE', (0, 0), (-1, 0), 10),
                    ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                    ('BACKGROUND', (0, 1), (-1, -2), colors.beige),
                    ('GRID', (0, 0), (-1, -1), 1, colors.grey),
                    ('BACKGROUND', (0, -1), (-1, -1), colors.HexColor('#FFD166')),
                    ('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold'),
                ]))
                
                story.append(table)
            else:
                story.append(Paragraph('No hay registros de tiempo para esta tarea.', styles['Normal']))
            
            story.append(Spacer(1, 0.4*inch))
        
        conn.close()
        
        # Generar PDF
        doc.build(story)
        
        return FileResponse(filename, filename=filename, media_type='application/pdf')
        
    except Exception as e:
        print(f"Error generando PDF: {e}")
        raise HTTPException(status_code=500, detail=str(e))

@app.get('/api/reports/date/excel')
async def generate_date_excel_report(
    from_date: str = Query(..., alias='from'),
    to_date: str = Query(..., alias='to'),
    user_id: Optional[int] = None,
    status: Optional[str] = None
):
    """Generar informe en Excel por rango de fechas"""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
        
        conn = get_db()
        
        query = '''
            SELECT t.*, u.name as user_name
            FROM tasks t
            LEFT JOIN users u ON t.user_id = u.id
            WHERE DATE(t.created_at) BETWEEN ? AND ?
        '''
        params = [from_date, to_date]
        
        if user_id:
            query += ' AND t.user_id = ?'
            params.append(int(user_id))
        
        if status:
            query += ' AND t.status = ?'
            params.append(status)
        
        query += ' ORDER BY t.created_at'
        
        tasks = conn.execute(query, params).fetchall()
        
        if not tasks:
            conn.close()
            raise HTTPException(status_code=404, detail='No se encontraron tareas en ese rango de fechas')
        
        wb = Workbook()
        wb.remove(wb.active)
        
        # Crear una hoja por cada tarea
        for task in tasks:
            task_dict = dict(task)
            sheet_name = f"{task_dict['task_number']} - {task_dict['name'][:25]}"
            ws = wb.create_sheet(sheet_name)
            
            # Título
            ws['A1'] = f"Tarea #{task_dict['task_number']}: {task_dict['name']}"
            ws['A1'].font = Font(size=14, bold=True, color='EF8354')
            ws.merge_cells('A1:E1')
            
            # Información
            ws['A2'] = f"Asignado a: {task_dict['user_name']}"
            ws['A3'] = f"Estado: {task_dict['status']}"
            ws['A4'] = f"Creada: {task_dict['created_at']}"
            ws['A5'] = f"Tiempo máximo: {task_dict['max_time_minutes']} minutos"
            
            if task_dict['max_date']:
                ws['A6'] = f"Fecha límite: {task_dict['max_date']}"
            
            # Encabezados
            row = 8
            ws[f'A{row}'] = 'Fecha/Hora Inicio'
            ws[f'B{row}'] = 'Fecha/Hora Fin'
            ws[f'C{row}'] = 'Duración (minutos)'
            ws[f'D{row}'] = 'Comentario'
            
            for col in ['A', 'B', 'C', 'D']:
                cell = ws[f'{col}{row}']
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color='4F5D75', end_color='4F5D75', fill_type='solid')
                cell.alignment = Alignment(horizontal='center')
            
            # Registros de tiempo
            times = conn.execute('''
                SELECT * FROM time_entries
                WHERE task_id = ?
                ORDER BY start_time
            ''', (task_dict['id'],)).fetchall()
            
            row += 1
            total_minutes = 0
            
            for time_entry in times:
                time_dict = dict(time_entry)
                ws[f'A{row}'] = time_dict['start_time']
                ws[f'B{row}'] = time_dict['end_time'] if time_dict['end_time'] else 'En progreso'
                
                if time_dict['duration_minutes']:
                    ws[f'C{row}'] = time_dict['duration_minutes']
                    total_minutes += time_dict['duration_minutes']
                else:
                    ws[f'C{row}'] = '-'
                
                ws[f'D{row}'] = time_dict['comment'] if time_dict['comment'] else '-'
                row += 1
            
            # Total
            row += 1
            ws[f'A{row}'] = 'TOTAL'
            ws[f'A{row}'].font = Font(bold=True)
            ws[f'C{row}'] = total_minutes
            ws[f'C{row}'].font = Font(bold=True)
            
            # Ajustar anchos
            ws.column_dimensions['A'].width = 20
            ws.column_dimensions['B'].width = 20
            ws.column_dimensions['C'].width = 20
            ws.column_dimensions['D'].width = 40
        
        conn.close()
        
        user_suffix = f'_usuario{user_id}' if user_id else ''
        status_suffix = f'_estado{status.replace(" ", "")}' if status else ''
        filename = f'informe_fechas_{from_date}_a_{to_date}{user_suffix}{status_suffix}.xlsx'
        wb.save(filename)
        
        return FileResponse(
            filename,
            filename=filename,
            media_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        
    except Exception as e:
        print(f"Error generando Excel por fechas: {e}")
        raise HTTPException(status_code=500, detail=str(e))

@app.get('/api/reports/date/pdf')
async def generate_date_pdf_report(
    from_date: str = Query(..., alias='from'),
    to_date: str = Query(..., alias='to'),
    user_id: Optional[int] = None,
    status: Optional[str] = None
):
    """Generar informe en PDF por rango de fechas"""
    try:
        from reportlab.lib.pagesizes import letter
        from reportlab.lib import colors
        from reportlab.lib.units import inch
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.lib.enums import TA_CENTER
        
        conn = get_db()
        
        query = '''
            SELECT t.*, u.name as user_name
            FROM tasks t
            LEFT JOIN users u ON t.user_id = u.id
            WHERE DATE(t.created_at) BETWEEN ? AND ?
        '''
        params = [from_date, to_date]
        
        if user_id:
            query += ' AND t.user_id = ?'
            params.append(int(user_id))
        
        if status:
            query += ' AND t.status = ?'
            params.append(status)
        
        query += ' ORDER BY t.created_at'
        
        tasks = conn.execute(query, params).fetchall()
        
        if not tasks:
            conn.close()
            raise HTTPException(status_code=404, detail='No se encontraron tareas en ese rango de fechas')
        
        user_suffix = f'_usuario{user_id}' if user_id else ''
        status_suffix = f'_estado{status.replace(" ", "")}' if status else ''
        filename = f'informe_fechas_{from_date}_a_{to_date}{user_suffix}{status_suffix}.pdf'
        doc = SimpleDocTemplate(filename, pagesize=letter)
        story = []
        styles = getSampleStyleSheet()
        
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=16,
            textColor=colors.HexColor('#EF8354'),
            spaceAfter=12,
            alignment=TA_CENTER
        )
        
        heading_style = ParagraphStyle(
            'CustomHeading',
            parent=styles['Heading2'],
            fontSize=14,
            textColor=colors.HexColor('#4F5D75'),
            spaceAfter=10
        )
        
        story.append(Paragraph('INFORME DE TAREAS POR FECHAS', title_style))
        story.append(Paragraph(f'Desde {from_date} hasta {to_date}', styles['Normal']))
        story.append(Spacer(1, 0.3*inch))
        
        for task in tasks:
            task_dict = dict(task)
            
            story.append(Paragraph(f"Tarea #{task_dict['task_number']}: {task_dict['name']}", heading_style))
            
            info_text = f"<b>Asignado a:</b> {task_dict['user_name']}<br/>"
            info_text += f"<b>Estado:</b> {task_dict['status']}<br/>"
            info_text += f"<b>Creada:</b> {task_dict['created_at']}<br/>"
            info_text += f"<b>Tiempo máximo:</b> {task_dict['max_time_minutes']} minutos<br/>"
            if task_dict['max_date']:
                info_text += f"<b>Fecha límite:</b> {task_dict['max_date']}<br/>"
            
            story.append(Paragraph(info_text, styles['Normal']))
            story.append(Spacer(1, 0.2*inch))
            
            times = conn.execute('''
                SELECT * FROM time_entries
                WHERE task_id = ?
                ORDER BY start_time
            ''', (task_dict['id'],)).fetchall()
            
            if times:
                data = [['Inicio', 'Fin', 'Duración', 'Comentario']]
                total_minutes = 0
                
                for time_entry in times:
                    time_dict = dict(time_entry)
                    data.append([
                        time_dict['start_time'],
                        time_dict['end_time'] if time_dict['end_time'] else 'En progreso',
                        str(time_dict['duration_minutes']) if time_dict['duration_minutes'] else '-',
                        time_dict['comment'] if time_dict['comment'] else '-'
                    ])
                    
                    if time_dict['duration_minutes']:
                        total_minutes += time_dict['duration_minutes']
                
                if total_minutes > 0:
                    data.append(['', '', str(total_minutes), 'TOTAL'])
                
                table = Table(data, colWidths=[1.5*inch, 1.5*inch, 1*inch, 3*inch])
                table.setStyle(TableStyle([
                    ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#4F5D75')),
                    ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                    ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                    ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                    ('FONTSIZE', (0, 0), (-1, 0), 10),
                    ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                    ('BACKGROUND', (0, 1), (-1, -2), colors.beige),
                    ('GRID', (0, 0), (-1, -1), 1, colors.grey),
                    ('BACKGROUND', (0, -1), (-1, -1), colors.HexColor('#FFD166')),
                    ('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold'),
                ]))
                
                story.append(table)
            else:
                story.append(Paragraph('No hay registros de tiempo.', styles['Normal']))
            
            story.append(Spacer(1, 0.4*inch))
        
        conn.close()
        doc.build(story)
        
        return FileResponse(filename, filename=filename, media_type='application/pdf')
        
    except Exception as e:
        print(f"Error generando PDF por fechas: {e}")
        raise HTTPException(status_code=500, detail=str(e))

@app.get('/api/reports/pending/excel')
async def generate_pending_excel_report(user_id: Optional[int] = None, status: Optional[str] = None):
    """Generar informe en Excel de tareas pendientes"""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
        
        conn = get_db()
        
        query = '''
            SELECT t.*, u.name as user_name
            FROM tasks t
            LEFT JOIN users u ON t.user_id = u.id
            WHERE t.status != 'Terminado'
        '''
        params = []
        
        if user_id:
            query += ' AND t.user_id = ?'
            params.append(int(user_id))
        
        if status:
            query += ' AND t.status = ?'
            params.append(status)
        
        query += ' ORDER BY t.task_number'
        
        tasks = conn.execute(query, params).fetchall()
        
        if not tasks:
            conn.close()
            raise HTTPException(status_code=404, detail='No se encontraron tareas pendientes')
        
        wb = Workbook()
        wb.remove(wb.active)
        
        for task in tasks:
            task_dict = dict(task)
            sheet_name = f"{task_dict['task_number']} - {task_dict['name'][:25]}"
            ws = wb.create_sheet(sheet_name)
            
            ws['A1'] = f"Tarea #{task_dict['task_number']}: {task_dict['name']}"
            ws['A1'].font = Font(size=14, bold=True, color='EF8354')
            ws.merge_cells('A1:E1')
            
            ws['A2'] = f"Asignado a: {task_dict['user_name']}"
            ws['A3'] = f"Estado: {task_dict['status']}"
            ws['A4'] = f"Tiempo máximo: {task_dict['max_time_minutes']} minutos"
            
            if task_dict['max_date']:
                ws['A5'] = f"Fecha límite: {task_dict['max_date']}"
            
            row = 7
            ws[f'A{row}'] = 'Fecha/Hora Inicio'
            ws[f'B{row}'] = 'Fecha/Hora Fin'
            ws[f'C{row}'] = 'Duración (minutos)'
            ws[f'D{row}'] = 'Comentario'
            
            for col in ['A', 'B', 'C', 'D']:
                cell = ws[f'{col}{row}']
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color='4F5D75', end_color='4F5D75', fill_type='solid')
                cell.alignment = Alignment(horizontal='center')
            
            times = conn.execute('''
                SELECT * FROM time_entries
                WHERE task_id = ?
                ORDER BY start_time
            ''', (task_dict['id'],)).fetchall()
            
            row += 1
            total_minutes = 0
            
            for time_entry in times:
                time_dict = dict(time_entry)
                ws[f'A{row}'] = time_dict['start_time']
                ws[f'B{row}'] = time_dict['end_time'] if time_dict['end_time'] else 'En progreso'
                
                if time_dict['duration_minutes']:
                    ws[f'C{row}'] = time_dict['duration_minutes']
                    total_minutes += time_dict['duration_minutes']
                else:
                    ws[f'C{row}'] = '-'
                
                ws[f'D{row}'] = time_dict['comment'] if time_dict['comment'] else '-'
                row += 1
            
            row += 1
            ws[f'A{row}'] = 'TOTAL'
            ws[f'A{row}'].font = Font(bold=True)
            ws[f'C{row}'] = total_minutes
            ws[f'C{row}'].font = Font(bold=True)
            
            ws.column_dimensions['A'].width = 20
            ws.column_dimensions['B'].width = 20
            ws.column_dimensions['C'].width = 20
            ws.column_dimensions['D'].width = 40
        
        conn.close()
        
        user_suffix = f'_usuario{user_id}' if user_id else ''
        status_suffix = f'_estado{status.replace(" ", "")}' if status else ''
        filename = f'informe_pendientes{user_suffix}{status_suffix}.xlsx'
        wb.save(filename)
        
        return FileResponse(
            filename,
            filename=filename,
            media_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        
    except Exception as e:
        print(f"Error generando Excel de pendientes: {e}")
        raise HTTPException(status_code=500, detail=str(e))

@app.get('/api/reports/pending/pdf')
async def generate_pending_pdf_report(user_id: Optional[int] = None, status: Optional[str] = None):
    """Generar informe en PDF de tareas pendientes"""
    try:
        from reportlab.lib.pagesizes import letter
        from reportlab.lib import colors
        from reportlab.lib.units import inch
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.lib.enums import TA_CENTER
        
        conn = get_db()
        
        query = '''
            SELECT t.*, u.name as user_name
            FROM tasks t
            LEFT JOIN users u ON t.user_id = u.id
            WHERE t.status != 'Terminado'
        '''
        params = []
        
        if user_id:
            query += ' AND t.user_id = ?'
            params.append(int(user_id))
        
        if status:
            query += ' AND t.status = ?'
            params.append(status)
        
        query += ' ORDER BY t.task_number'
        
        tasks = conn.execute(query, params).fetchall()
        
        if not tasks:
            conn.close()
            raise HTTPException(status_code=404, detail='No se encontraron tareas pendientes')
        
        user_suffix = f'_usuario{user_id}' if user_id else ''
        status_suffix = f'_estado{status.replace(" ", "")}' if status else ''
        filename = f'informe_pendientes{user_suffix}{status_suffix}.pdf'
        doc = SimpleDocTemplate(filename, pagesize=letter)
        story = []
        styles = getSampleStyleSheet()
        
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=16,
            textColor=colors.HexColor('#EF8354'),
            spaceAfter=12,
            alignment=TA_CENTER
        )
        
        heading_style = ParagraphStyle(
            'CustomHeading',
            parent=styles['Heading2'],
            fontSize=14,
            textColor=colors.HexColor('#4F5D75'),
            spaceAfter=10
        )
        
        story.append(Paragraph('INFORME DE TAREAS PENDIENTES', title_style))
        story.append(Spacer(1, 0.3*inch))
        
        for task in tasks:
            task_dict = dict(task)
            
            story.append(Paragraph(f"Tarea #{task_dict['task_number']}: {task_dict['name']}", heading_style))
            
            info_text = f"<b>Asignado a:</b> {task_dict['user_name']}<br/>"
            info_text += f"<b>Estado:</b> {task_dict['status']}<br/>"
            info_text += f"<b>Tiempo máximo:</b> {task_dict['max_time_minutes']} minutos<br/>"
            if task_dict['max_date']:
                info_text += f"<b>Fecha límite:</b> {task_dict['max_date']}<br/>"
            
            story.append(Paragraph(info_text, styles['Normal']))
            story.append(Spacer(1, 0.2*inch))
            
            times = conn.execute('''
                SELECT * FROM time_entries
                WHERE task_id = ?
                ORDER BY start_time
            ''', (task_dict['id'],)).fetchall()
            
            if times:
                data = [['Inicio', 'Fin', 'Duración', 'Comentario']]
                total_minutes = 0
                
                for time_entry in times:
                    time_dict = dict(time_entry)
                    data.append([
                        time_dict['start_time'],
                        time_dict['end_time'] if time_dict['end_time'] else 'En progreso',
                        str(time_dict['duration_minutes']) if time_dict['duration_minutes'] else '-',
                        time_dict['comment'] if time_dict['comment'] else '-'
                    ])
                    
                    if time_dict['duration_minutes']:
                        total_minutes += time_dict['duration_minutes']
                
                if total_minutes > 0:
                    data.append(['', '', str(total_minutes), 'TOTAL'])
                
                table = Table(data, colWidths=[1.5*inch, 1.5*inch, 1*inch, 3*inch])
                table.setStyle(TableStyle([
                    ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#4F5D75')),
                    ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                    ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                    ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                    ('FONTSIZE', (0, 0), (-1, 0), 10),
                    ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                    ('BACKGROUND', (0, 1), (-1, -2), colors.beige),
                    ('GRID', (0, 0), (-1, -1), 1, colors.grey),
                    ('BACKGROUND', (0, -1), (-1, -1), colors.HexColor('#FFD166')),
                    ('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold'),
                ]))
                
                story.append(table)
            else:
                story.append(Paragraph('No hay registros de tiempo.', styles['Normal']))
            
            story.append(Spacer(1, 0.4*inch))
        
        conn.close()
        doc.build(story)
        
        return FileResponse(filename, filename=filename, media_type='application/pdf')
        
    except Exception as e:
        print(f"Error generando PDF de pendientes: {e}")
        raise HTTPException(status_code=500, detail=str(e))

# ==================== INFORME DE REGISTROS DE TIEMPO ====================

@app.get('/api/reports/timeentries/excel')
async def generate_timeentries_excel_report(
    from_date: str = Query(..., alias='from'),
    to_date: str = Query(..., alias='to'),
    user_id: Optional[int] = None
):
    """Generar informe de registros de tiempo en Excel"""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
        from datetime import datetime, timedelta
        
        conn = get_db()
        
        # Query para obtener registros de tiempo con información de tarea y usuario
        query = '''
            SELECT 
                te.id,
                te.start_time,
                te.end_time,
                te.duration_minutes,
                te.comment,
                t.name as task_name,
                t.task_number,
                u.name as user_name,
                t.user_id
            FROM time_entries te
            JOIN tasks t ON te.task_id = t.id
            LEFT JOIN users u ON t.user_id = u.id
            WHERE DATE(te.start_time) >= ? AND DATE(te.start_time) <= ?
        '''
        params = [from_date, to_date]
        
        if user_id:
            query += ' AND t.user_id = ?'
            params.append(int(user_id))
        
        query += ' ORDER BY te.start_time'
        
        entries = conn.execute(query, params).fetchall()
        
        if not entries:
            conn.close()
            raise HTTPException(status_code=404, detail='No se encontraron registros en ese rango de fechas')
        
        wb = Workbook()
        ws = wb.active
        ws.title = "Registros de Tiempo"
        
        # Título
        ws['A1'] = 'INFORME DE REGISTROS DE TIEMPO'
        ws['A1'].font = Font(size=16, bold=True, color='EF8354')
        ws.merge_cells('A1:F1')
        ws['A1'].alignment = Alignment(horizontal='center')
        
        ws['A2'] = f'Periodo: {from_date} a {to_date}'
        ws.merge_cells('A2:F2')
        ws['A2'].alignment = Alignment(horizontal='center')
        
        # Encabezados
        row = 4
        headers = ['Fecha/Hora Inicio', 'Fecha/Hora Fin', 'Duración (min)', 'Tarea', 'Usuario', 'Comentario']
        for col_idx, header in enumerate(headers, start=1):
            cell = ws.cell(row=row, column=col_idx, value=header)
            cell.font = Font(bold=True, color='FFFFFF')
            cell.fill = PatternFill(start_color='4F5D75', end_color='4F5D75', fill_type='solid')
            cell.alignment = Alignment(horizontal='center')
        
        # Datos
        row += 1
        total_minutes = 0
        
        for entry in entries:
            entry_dict = dict(entry)
            
            # Fecha/Hora Inicio
            ws.cell(row=row, column=1, value=entry_dict['start_time'])
            
            # Fecha/Hora Fin
            end_time_cell = ws.cell(row=row, column=2)
            if entry_dict['end_time']:
                end_time_cell.value = entry_dict['end_time']
            else:
                # Calcular fin del día a las 20:00
                start_dt = datetime.fromisoformat(entry_dict['start_time'].replace('Z', '+00:00'))
                end_of_day = start_dt.replace(hour=20, minute=0, second=0, microsecond=0)
                end_time_cell.value = end_of_day.strftime('%Y-%m-%d %H:%M:%S') + ' *'
                # Fondo rojo, texto blanco
                end_time_cell.fill = PatternFill(start_color='FF0000', end_color='FF0000', fill_type='solid')
                end_time_cell.font = Font(color='FFFFFF', bold=True)
            
            # Duración
            duration_minutes = entry_dict['duration_minutes']
            if not duration_minutes and entry_dict['start_time']:
                # Calcular duración hasta las 20:00 del mismo día
                start_dt = datetime.fromisoformat(entry_dict['start_time'].replace('Z', '+00:00'))
                end_of_day = start_dt.replace(hour=20, minute=0, second=0, microsecond=0)
                duration_minutes = int((end_of_day - start_dt).total_seconds() / 60)
            
            ws.cell(row=row, column=3, value=duration_minutes if duration_minutes else 0)
            if duration_minutes:
                total_minutes += duration_minutes
            
            # Tarea
            ws.cell(row=row, column=4, value=f"#{entry_dict['task_number']}: {entry_dict['task_name']}")
            
            # Usuario
            ws.cell(row=row, column=5, value=entry_dict['user_name'])
            
            # Comentario
            ws.cell(row=row, column=6, value=entry_dict['comment'] if entry_dict['comment'] else '-')
            
            row += 1
        
        # Total
        row += 1
        total_cell = ws.cell(row=row, column=2, value='TOTAL:')
        total_cell.font = Font(bold=True, size=12)
        total_cell.alignment = Alignment(horizontal='right')
        
        total_value_cell = ws.cell(row=row, column=3, value=total_minutes)
        total_value_cell.font = Font(bold=True, size=12)
        total_value_cell.fill = PatternFill(start_color='FFD166', end_color='FFD166', fill_type='solid')
        
        # Ajustar anchos de columna
        ws.column_dimensions['A'].width = 20
        ws.column_dimensions['B'].width = 22
        ws.column_dimensions['C'].width = 15
        ws.column_dimensions['D'].width = 35
        ws.column_dimensions['E'].width = 20
        ws.column_dimensions['F'].width = 40
        
        conn.close()
        
        # Guardar archivo
        user_suffix = f'_usuario{user_id}' if user_id else ''
        filename = f'informe_registros_{from_date}_a_{to_date}{user_suffix}.xlsx'
        wb.save(filename)
        
        return FileResponse(
            filename,
            filename=filename,
            media_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        
    except Exception as e:
        print(f"Error generando Excel de registros: {e}")
        raise HTTPException(status_code=500, detail=str(e))

@app.get('/api/reports/timeentries/pdf')
async def generate_timeentries_pdf_report(
    from_date: str = Query(..., alias='from'),
    to_date: str = Query(..., alias='to'),
    user_id: Optional[int] = None
):
    """Generar informe de registros de tiempo en PDF"""
    try:
        from reportlab.lib.pagesizes import letter, landscape
        from reportlab.lib import colors
        from reportlab.lib.units import inch
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.lib.enums import TA_CENTER
        from datetime import datetime, timedelta
        
        conn = get_db()
        
        # Query para obtener registros
        query = '''
            SELECT 
                te.id,
                te.start_time,
                te.end_time,
                te.duration_minutes,
                te.comment,
                t.name as task_name,
                t.task_number,
                u.name as user_name,
                t.user_id
            FROM time_entries te
            JOIN tasks t ON te.task_id = t.id
            LEFT JOIN users u ON t.user_id = u.id
            WHERE DATE(te.start_time) >= ? AND DATE(te.start_time) <= ?
        '''
        params = [from_date, to_date]
        
        if user_id:
            query += ' AND t.user_id = ?'
            params.append(int(user_id))
        
        query += ' ORDER BY te.start_time'
        
        entries = conn.execute(query, params).fetchall()
        
        if not entries:
            conn.close()
            raise HTTPException(status_code=404, detail='No se encontraron registros en ese rango de fechas')
        
        user_suffix = f'_usuario{user_id}' if user_id else ''
        filename = f'informe_registros_{from_date}_a_{to_date}{user_suffix}.pdf'
        
        # Usar landscape para más espacio horizontal
        doc = SimpleDocTemplate(filename, pagesize=landscape(letter))
        story = []
        styles = getSampleStyleSheet()
        
        # Estilos personalizados
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=18,
            textColor=colors.HexColor('#EF8354'),
            spaceAfter=12,
            alignment=TA_CENTER
        )
        
        # Título
        story.append(Paragraph('INFORME DE REGISTROS DE TIEMPO', title_style))
        story.append(Paragraph(f'Periodo: {from_date} a {to_date}', styles['Normal']))
        story.append(Spacer(1, 0.3*inch))
        
        # Preparar datos de la tabla
        data = [['Inicio', 'Fin', 'Duración', 'Tarea', 'Usuario', 'Comentario']]
        total_minutes = 0
        
        for entry in entries:
            entry_dict = dict(entry)
            
            # Inicio
            start_time = entry_dict['start_time']
            
            # Fin
            if entry_dict['end_time']:
                end_time = entry_dict['end_time']
            else:
                # Calcular fin del día a las 20:00
                start_dt = datetime.fromisoformat(entry_dict['start_time'].replace('Z', '+00:00'))
                end_of_day = start_dt.replace(hour=20, minute=0, second=0, microsecond=0)
                end_time = end_of_day.strftime('%Y-%m-%d %H:%M:%S') + ' *'
            
            # Duración
            duration_minutes = entry_dict['duration_minutes']
            if not duration_minutes and entry_dict['start_time']:
                start_dt = datetime.fromisoformat(entry_dict['start_time'].replace('Z', '+00:00'))
                end_of_day = start_dt.replace(hour=20, minute=0, second=0, microsecond=0)
                duration_minutes = int((end_of_day - start_dt).total_seconds() / 60)
            
            if duration_minutes:
                total_minutes += duration_minutes
            
            # Tarea
            task = f"#{entry_dict['task_number']}: {entry_dict['task_name'][:30]}"
            
            # Usuario
            user = entry_dict['user_name'][:15] if entry_dict['user_name'] else '-'
            
            # Comentario
            comment = entry_dict['comment'][:35] if entry_dict['comment'] else '-'
            
            data.append([
                start_time,
                end_time,
                f"{duration_minutes} min" if duration_minutes else '-',
                task,
                user,
                comment
            ])
        
        # Fila de total
        data.append(['', 'TOTAL:', f'{total_minutes} min', '', '', ''])
        
        # Crear tabla
        table = Table(data, colWidths=[1.3*inch, 1.4*inch, 0.8*inch, 2*inch, 1.2*inch, 2*inch])
        
        # Estilo de la tabla
        table_style = [
            # Encabezado
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#4F5D75')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 10),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            # Datos
            ('BACKGROUND', (0, 1), (-1, -2), colors.beige),
            ('GRID', (0, 0), (-1, -1), 1, colors.grey),
            ('FONTSIZE', (0, 1), (-1, -1), 8),
            # Total
            ('BACKGROUND', (0, -1), (-1, -1), colors.HexColor('#FFD166')),
            ('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold'),
            ('FONTSIZE', (0, -1), (-1, -1), 11),
        ]
        
        # Añadir fondo rojo para celdas con asterisco (fin no registrado)
        for row_idx, row in enumerate(data[1:], start=1):  # Empezar desde fila 1 (después del header)
            if row_idx < len(data) - 1:  # No aplicar al total
                if '*' in str(row[1]):  # Si la columna Fin tiene asterisco
                    table_style.append(('BACKGROUND', (1, row_idx), (1, row_idx), colors.red))
                    table_style.append(('TEXTCOLOR', (1, row_idx), (1, row_idx), colors.white))
                    table_style.append(('FONTNAME', (1, row_idx), (1, row_idx), 'Helvetica-Bold'))
        
        table.setStyle(TableStyle(table_style))
        story.append(table)
        
        # Nota al pie
        story.append(Spacer(1, 0.2*inch))
        note_style = ParagraphStyle('Note', parent=styles['Normal'], fontSize=9, textColor=colors.grey)
        story.append(Paragraph('* Registros sin hora de fin: se calcula duración hasta las 20:00 del mismo día', note_style))
        
        conn.close()
        
        # Generar PDF
        doc.build(story)
        
        return FileResponse(filename, filename=filename, media_type='application/pdf')
        
    except Exception as e:
        print(f"Error generando PDF de registros: {e}")
        raise HTTPException(status_code=500, detail=str(e))

# ==================== GESTIÓN DE REGISTROS DE TIEMPO ====================

@app.get('/api/timeentries/list')
async def list_time_entries(
    from_date: str = Query(...),
    to_date: str = Query(...),
    user_id: Optional[int] = None,
    has_end: Optional[str] = None,  # 'yes', 'no', 'all'
    status: Optional[str] = None
):
    """Listar registros de tiempo con filtros avanzados"""
    try:
        conn = get_db()
        
        query = '''
            SELECT 
                te.id,
                te.task_id,
                te.start_time,
                te.end_time,
                te.duration_minutes,
                te.comment,
                t.name as task_name,
                t.task_number,
                t.status as task_status,
                u.name as user_name,
                t.user_id
            FROM time_entries te
            JOIN tasks t ON te.task_id = t.id
            LEFT JOIN users u ON t.user_id = u.id
            WHERE DATE(te.start_time) >= ? AND DATE(te.start_time) <= ?
        '''
        params = [from_date, to_date]
        
        # Filtro por usuario
        if user_id:
            query += ' AND t.user_id = ?'
            params.append(int(user_id))
        
        # Filtro por si tiene fin o no
        if has_end == 'yes':
            query += ' AND te.end_time IS NOT NULL'
        elif has_end == 'no':
            query += ' AND te.end_time IS NULL'
        # 'all' no añade filtro
        
        # Filtro por estado de tarea
        if status:
            query += ' AND t.status = ?'
            params.append(status)
        
        query += ' ORDER BY te.start_time DESC'
        
        entries = conn.execute(query, params).fetchall()
        conn.close()
        
        return [dict(entry) for entry in entries]
        
    except Exception as e:
        print(f"Error listando registros: {e}")
        raise HTTPException(status_code=500, detail=str(e))

@app.get('/api/timeentries/export/excel')
async def export_time_entries_excel(
    from_date: str = Query(...),
    to_date: str = Query(...),
    user_id: Optional[int] = None,
    has_end: Optional[str] = None,
    status: Optional[str] = None
):
    """Exportar registros de tiempo a Excel con filtros"""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
        from datetime import datetime
        
        conn = get_db()
        
        # Misma query que list
        query = '''
            SELECT 
                te.id,
                te.task_id,
                te.start_time,
                te.end_time,
                te.duration_minutes,
                te.comment,
                t.name as task_name,
                t.task_number,
                t.status as task_status,
                u.name as user_name,
                t.user_id
            FROM time_entries te
            JOIN tasks t ON te.task_id = t.id
            LEFT JOIN users u ON t.user_id = u.id
            WHERE DATE(te.start_time) >= ? AND DATE(te.start_time) <= ?
        '''
        params = [from_date, to_date]
        
        if user_id:
            query += ' AND t.user_id = ?'
            params.append(int(user_id))
        
        if has_end == 'yes':
            query += ' AND te.end_time IS NOT NULL'
        elif has_end == 'no':
            query += ' AND te.end_time IS NULL'
        
        if status:
            query += ' AND t.status = ?'
            params.append(status)
        
        query += ' ORDER BY te.start_time DESC'
        
        entries = conn.execute(query, params).fetchall()
        
        if not entries:
            conn.close()
            raise HTTPException(status_code=404, detail='No se encontraron registros con esos filtros')
        
        wb = Workbook()
        ws = wb.active
        ws.title = "Registros"
        
        # Título
        ws['A1'] = 'LISTADO DE REGISTROS DE TIEMPO'
        ws['A1'].font = Font(size=16, bold=True, color='EF8354')
        ws.merge_cells('A1:H1')
        ws['A1'].alignment = Alignment(horizontal='center')
        
        # Filtros aplicados
        filter_info = f"Periodo: {from_date} a {to_date}"
        if user_id:
            user_name = entries[0]['user_name'] if entries else 'N/A'
            filter_info += f" | Usuario: {user_name}"
        if has_end == 'yes':
            filter_info += " | Con fecha fin"
        elif has_end == 'no':
            filter_info += " | Sin fecha fin"
        if status:
            filter_info += f" | Estado: {status}"
        
        ws['A2'] = filter_info
        ws.merge_cells('A2:H2')
        ws['A2'].alignment = Alignment(horizontal='center')
        
        # Encabezados
        row = 4
        headers = ['ID', 'Inicio', 'Fin', 'Duración', 'Tarea', 'Estado', 'Usuario', 'Comentario']
        for col_idx, header in enumerate(headers, start=1):
            cell = ws.cell(row=row, column=col_idx, value=header)
            cell.font = Font(bold=True, color='FFFFFF')
            cell.fill = PatternFill(start_color='4F5D75', end_color='4F5D75', fill_type='solid')
            cell.alignment = Alignment(horizontal='center')
        
        # Datos
        row += 1
        total_minutes = 0
        
        for entry in entries:
            entry_dict = dict(entry)
            
            ws.cell(row=row, column=1, value=entry_dict['id'])
            ws.cell(row=row, column=2, value=entry_dict['start_time'])
            
            # Fin
            end_cell = ws.cell(row=row, column=3)
            if entry_dict['end_time']:
                end_cell.value = entry_dict['end_time']
            else:
                start_dt = datetime.fromisoformat(entry_dict['start_time'].replace('Z', '+00:00'))
                end_of_day = start_dt.replace(hour=20, minute=0, second=0)
                end_cell.value = end_of_day.strftime('%Y-%m-%d %H:%M:%S') + ' *'
                end_cell.fill = PatternFill(start_color='FF0000', end_color='FF0000', fill_type='solid')
                end_cell.font = Font(color='FFFFFF', bold=True)
            
            # Duración
            duration = entry_dict['duration_minutes']
            if not duration and entry_dict['start_time']:
                start_dt = datetime.fromisoformat(entry_dict['start_time'].replace('Z', '+00:00'))
                end_of_day = start_dt.replace(hour=20, minute=0, second=0)
                duration = int((end_of_day - start_dt).total_seconds() / 60)
            
            ws.cell(row=row, column=4, value=duration if duration else 0)
            if duration:
                total_minutes += duration
            
            ws.cell(row=row, column=5, value=f"#{entry_dict['task_number']}: {entry_dict['task_name']}")
            ws.cell(row=row, column=6, value=entry_dict['task_status'])
            ws.cell(row=row, column=7, value=entry_dict['user_name'])
            ws.cell(row=row, column=8, value=entry_dict['comment'] if entry_dict['comment'] else '-')
            
            row += 1
        
        # Total
        row += 1
        ws.cell(row=row, column=3, value='TOTAL:').font = Font(bold=True)
        ws.cell(row=row, column=4, value=total_minutes).font = Font(bold=True)
        ws.cell(row=row, column=4).fill = PatternFill(start_color='FFD166', end_color='FFD166', fill_type='solid')
        
        # Ajustar anchos
        ws.column_dimensions['A'].width = 8
        ws.column_dimensions['B'].width = 20
        ws.column_dimensions['C'].width = 22
        ws.column_dimensions['D'].width = 12
        ws.column_dimensions['E'].width = 35
        ws.column_dimensions['F'].width = 15
        ws.column_dimensions['G'].width = 20
        ws.column_dimensions['H'].width = 40
        
        conn.close()
        
        # Nombre de archivo
        filename = f'registros_{from_date}_a_{to_date}'
        if user_id:
            filename += f'_usuario{user_id}'
        if has_end == 'yes':
            filename += '_finalizados'
        elif has_end == 'no':
            filename += '_sinFinalizar'
        if status:
            filename += f'_estado{status.replace(" ", "")}'
        filename += '.xlsx'
        
        wb.save(filename)
        
        return FileResponse(
            filename,
            filename=filename,
            media_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        
    except Exception as e:
        print(f"Error exportando registros a Excel: {e}")
        raise HTTPException(status_code=500, detail=str(e))

@app.get('/api/timeentries/export/pdf')
async def export_time_entries_pdf(
    from_date: str = Query(...),
    to_date: str = Query(...),
    user_id: Optional[int] = None,
    has_end: Optional[str] = None,
    status: Optional[str] = None
):
    """Exportar registros de tiempo a PDF con filtros"""
    try:
        from reportlab.lib.pagesizes import letter, landscape
        from reportlab.lib import colors
        from reportlab.lib.units import inch
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.lib.enums import TA_CENTER
        from datetime import datetime
        
        conn = get_db()
        
        # Misma query
        query = '''
            SELECT 
                te.id,
                te.task_id,
                te.start_time,
                te.end_time,
                te.duration_minutes,
                te.comment,
                t.name as task_name,
                t.task_number,
                t.status as task_status,
                u.name as user_name,
                t.user_id
            FROM time_entries te
            JOIN tasks t ON te.task_id = t.id
            LEFT JOIN users u ON t.user_id = u.id
            WHERE DATE(te.start_time) >= ? AND DATE(te.start_time) <= ?
        '''
        params = [from_date, to_date]
        
        if user_id:
            query += ' AND t.user_id = ?'
            params.append(int(user_id))
        
        if has_end == 'yes':
            query += ' AND te.end_time IS NOT NULL'
        elif has_end == 'no':
            query += ' AND te.end_time IS NULL'
        
        if status:
            query += ' AND t.status = ?'
            params.append(status)
        
        query += ' ORDER BY te.start_time DESC'
        
        entries = conn.execute(query, params).fetchall()
        
        if not entries:
            conn.close()
            raise HTTPException(status_code=404, detail='No se encontraron registros con esos filtros')
        
        # Nombre de archivo
        filename = f'registros_{from_date}_a_{to_date}'
        if user_id:
            filename += f'_usuario{user_id}'
        if has_end == 'yes':
            filename += '_finalizados'
        elif has_end == 'no':
            filename += '_sinFinalizar'
        if status:
            filename += f'_estado{status.replace(" ", "")}'
        filename += '.pdf'
        
        doc = SimpleDocTemplate(filename, pagesize=landscape(letter))
        story = []
        styles = getSampleStyleSheet()
        
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=18,
            textColor=colors.HexColor('#EF8354'),
            spaceAfter=12,
            alignment=TA_CENTER
        )
        
        story.append(Paragraph('LISTADO DE REGISTROS DE TIEMPO', title_style))
        
        # Filtros
        filter_info = f"Periodo: {from_date} a {to_date}"
        if user_id:
            filter_info += f" | Usuario: {entries[0]['user_name']}"
        if has_end == 'yes':
            filter_info += " | Con fecha fin"
        elif has_end == 'no':
            filter_info += " | Sin fecha fin"
        if status:
            filter_info += f" | Estado: {status}"
        
        story.append(Paragraph(filter_info, styles['Normal']))
        story.append(Spacer(1, 0.3*inch))
        
        # Tabla
        data = [['ID', 'Inicio', 'Fin', 'Dur', 'Tarea', 'Estado', 'Usuario', 'Comentario']]
        total_minutes = 0
        
        for entry in entries:
            entry_dict = dict(entry)
            
            # Fin
            if entry_dict['end_time']:
                end_time = entry_dict['end_time']
            else:
                start_dt = datetime.fromisoformat(entry_dict['start_time'].replace('Z', '+00:00'))
                end_of_day = start_dt.replace(hour=20, minute=0, second=0)
                end_time = end_of_day.strftime('%Y-%m-%d %H:%M:%S') + ' *'
            
            # Duración
            duration = entry_dict['duration_minutes']
            if not duration and entry_dict['start_time']:
                start_dt = datetime.fromisoformat(entry_dict['start_time'].replace('Z', '+00:00'))
                end_of_day = start_dt.replace(hour=20, minute=0, second=0)
                duration = int((end_of_day - start_dt).total_seconds() / 60)
            
            if duration:
                total_minutes += duration
            
            data.append([
                str(entry_dict['id']),
                entry_dict['start_time'],
                end_time,
                f"{duration}m" if duration else '-',
                f"#{entry_dict['task_number']}: {entry_dict['task_name'][:25]}",
                entry_dict['task_status'][:10],
                entry_dict['user_name'][:12] if entry_dict['user_name'] else '-',
                entry_dict['comment'][:30] if entry_dict['comment'] else '-'
            ])
        
        # Total
        data.append(['', '', 'TOTAL:', f'{total_minutes}m', '', '', '', ''])
        
        table = Table(data, colWidths=[0.4*inch, 1.3*inch, 1.4*inch, 0.6*inch, 2*inch, 0.9*inch, 1*inch, 1.8*inch])
        
        table_style = [
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#4F5D75')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 9),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -2), colors.beige),
            ('GRID', (0, 0), (-1, -1), 1, colors.grey),
            ('FONTSIZE', (0, 1), (-1, -1), 7),
            ('BACKGROUND', (0, -1), (-1, -1), colors.HexColor('#FFD166')),
            ('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold'),
        ]
        
        # Fondo rojo para registros sin fin
        for row_idx, row in enumerate(data[1:], start=1):
            if row_idx < len(data) - 1:
                if '*' in str(row[2]):
                    table_style.append(('BACKGROUND', (2, row_idx), (2, row_idx), colors.red))
                    table_style.append(('TEXTCOLOR', (2, row_idx), (2, row_idx), colors.white))
        
        table.setStyle(TableStyle(table_style))
        story.append(table)
        
        story.append(Spacer(1, 0.2*inch))
        note = ParagraphStyle('Note', parent=styles['Normal'], fontSize=8, textColor=colors.grey)
        story.append(Paragraph('* Registros sin hora de fin: duración calculada hasta 20:00', note))
        
        conn.close()
        
        doc.build(story)
        
        return FileResponse(filename, filename=filename, media_type='application/pdf')
        
    except Exception as e:
        print(f"Error exportando registros a PDF: {e}")
        raise HTTPException(status_code=500, detail=str(e))

# ==================== CRUD INDIVIDUAL DE REGISTROS ====================

@app.get('/api/timeentries/{entry_id}')
async def get_time_entry(entry_id: int):
    """Obtener un registro de tiempo específico"""
    try:
        conn = get_db()
        
        query = '''
            SELECT 
                te.id,
                te.task_id,
                te.start_time,
                te.end_time,
                te.duration_minutes,
                te.comment,
                t.name as task_name,
                t.task_number,
                t.status as task_status,
                u.name as user_name
            FROM time_entries te
            JOIN tasks t ON te.task_id = t.id
            LEFT JOIN users u ON t.user_id = u.id
            WHERE te.id = ?
        '''
        
        entry = conn.execute(query, (entry_id,)).fetchone()
        conn.close()
        
        if not entry:
            raise HTTPException(status_code=404, detail='Registro no encontrado')
        
        return dict(entry)
        
    except HTTPException:
        raise
    except Exception as e:
        print(f"Error obteniendo registro: {e}")
        raise HTTPException(status_code=500, detail=str(e))

@app.put('/api/timeentries/{entry_id}')
async def update_time_entry(entry_id: int, entry_data: dict):
    """Actualizar un registro de tiempo"""
    try:
        conn = get_db()
        
        # Verificar que existe
        existing = conn.execute('SELECT id FROM time_entries WHERE id = ?', (entry_id,)).fetchone()
        if not existing:
            conn.close()
            raise HTTPException(status_code=404, detail='Registro no encontrado')
        
        start_time = entry_data.get('start_time')
        end_time = entry_data.get('end_time')
        comment = entry_data.get('comment', '')
        
        if not start_time:
            conn.close()
            raise HTTPException(status_code=400, detail='La fecha de inicio es obligatoria')
        
        # Calcular duración si hay end_time
        duration_minutes = None
        if end_time:
            from datetime import datetime
            start_dt = datetime.fromisoformat(start_time.replace('Z', '+00:00'))
            end_dt = datetime.fromisoformat(end_time.replace('Z', '+00:00'))
            duration_minutes = int((end_dt - start_dt).total_seconds() / 60)
        
        # Actualizar registro
        conn.execute('''
            UPDATE time_entries 
            SET start_time = ?, end_time = ?, duration_minutes = ?, comment = ?
            WHERE id = ?
        ''', (start_time, end_time, duration_minutes, comment, entry_id))
        
        conn.commit()
        
        # Obtener registro actualizado
        updated = conn.execute('''
            SELECT 
                te.id,
                te.task_id,
                te.start_time,
                te.end_time,
                te.duration_minutes,
                te.comment,
                t.name as task_name,
                t.task_number
            FROM time_entries te
            JOIN tasks t ON te.task_id = t.id
            WHERE te.id = ?
        ''', (entry_id,)).fetchone()
        
        conn.close()
        
        return dict(updated)
        
    except HTTPException:
        raise
    except Exception as e:
        print(f"Error actualizando registro: {e}")
        raise HTTPException(status_code=500, detail=str(e))

@app.delete('/api/timeentries/{entry_id}')
async def delete_time_entry(entry_id: int):
    """Eliminar un registro de tiempo"""
    try:
        conn = get_db()
        
        # Verificar que existe
        existing = conn.execute('SELECT id FROM time_entries WHERE id = ?', (entry_id,)).fetchone()
        if not existing:
            conn.close()
            raise HTTPException(status_code=404, detail='Registro no encontrado')
        
        # Eliminar
        conn.execute('DELETE FROM time_entries WHERE id = ?', (entry_id,))
        conn.commit()
        conn.close()
        
        return {'message': 'Registro eliminado exitosamente'}
        
    except HTTPException:
        raise
    except Exception as e:
        print(f"Error eliminando registro: {e}")
        raise HTTPException(status_code=500, detail=str(e))
